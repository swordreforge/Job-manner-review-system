#!/usr/bin/env python3
"""
生成批量导入模板脚本
用于生成学生和岗位的 Excel 导入模板
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def create_student_template():
    """创建学生导入模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "学生数据"

    # 设置列宽
    column_widths = {
        'A': 15,  # 姓名
        'B': 12,  # 学历
        'C': 20,  # 专业
        'D': 12,  # 毕业年份
        'E': 30,  # 技能
        'F': 25,  # 证书
        'G': 20,  # 软技能
        'H': 30,  # 实习经历
        'I': 30,  # 项目经验
        'J': 25,  # 备注
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 定义样式
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 标题行
    headers = [
        '姓名* (必填)', '学历', '专业', '毕业年份', 
        '技能', '证书', '软技能', '实习经历', '项目经验', '备注'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        cell.value = header

    # 设置行高
    ws.row_dimensions[1].height = 30

    # 示例数据
    example_data = [
        ['张三', '本科', '计算机科学与技术', '2024', 'Golang,Python,MySQL,Redis', '无', '团队协作,沟通能力', '字节跳动实习', '电商系统开发', '优秀学生'],
        ['李四', '本科', '软件工程', '2024', 'Java,Spring,MySQL,Vue', '英语六级', '沟通能力,学习能力强', '阿里巴巴实习', '管理系统开发', '优秀学生'],
        ['王五', '硕士', '人工智能', '2025', 'Python,TensorFlow,PyTorch', '无', '团队协作,创新思维', '腾讯实习', 'AI模型训练', '优秀学生'],
    ]

    for row_num, data in enumerate(example_data, 2):
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.value = value

    # 设置数据行高
    for row_num in range(2, 5):
        ws.row_dimensions[row_num].height = 40

    # 添加说明工作表
    info_ws = wb.create_sheet("说明")
    info_column_widths = {'A': 40, 'B': 60}
    for col, width in info_column_widths.items():
        info_ws.column_dimensions[col].width = width

    info_headers = ['字段', '说明']
    for col_num, header in enumerate(info_headers, 1):
        cell = info_ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        cell.value = header

    info_ws.row_dimensions[1].height = 30

    info_data = [
        ['姓名*', '学生姓名，必填项'],
        ['学历', '如：高中、大专、本科、硕士、博士'],
        ['专业', '学生主修专业'],
        ['毕业年份', '毕业年份，格式为四位数字，如：2024'],
        ['技能', '使用逗号分隔，如：Golang,Python,MySQL'],
        ['证书', '已获得的证书，使用逗号分隔'],
        ['软技能', '如：团队协作、沟通能力、学习能力等'],
        ['实习经历', '简述实习经历和公司'],
        ['项目经验', '简述参与的项目和角色'],
        ['备注', '其他需要说明的信息'],
    ]

    for row_num, data in enumerate(info_data, 2):
        for col_num, value in enumerate(data, 1):
            cell = info_ws.cell(row=row_num, column=col_num)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.value = value
            info_ws.row_dimensions[row_num].height = 40

    # 保存文件
    output_path = os.path.join(os.path.dirname(__file__), '../template/student_import_template.xlsx')
    wb.save(output_path)
    print(f"✓ 学生导入模板已生成: {output_path}")

def create_job_template():
    """创建岗位导入模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "岗位数据"

    # 设置列宽
    column_widths = {
        'A': 20,  # 岗位名称
        'B': 30,  # 描述
        'C': 20,  # 公司
        'D': 12,  # 行业
        'E': 12,  # 类别
        'F': 15,  # 地点
        'G': 15,  # 薪资范围
        'H': 30,  # 技能要求
        'I': 25,  # 证书要求
        'J': 20,  # 软技能
        'K': 25,  # 岗位要求
        'L': 10,  # 成长潜力
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 定义样式
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 标题行
    headers = [
        '岗位名称* (必填)', '描述', '公司', '行业', '类别', '地点', 
        '薪资范围', '技能要求', '证书要求', '软技能', '岗位要求', '成长潜力'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        cell.value = header

    # 设置行高
    ws.row_dimensions[1].height = 30

    # 示例数据
    example_data = [
        ['Golang后端开发工程师', '负责公司后端服务开发和维护', '字节跳动', '技术', '开发', '北京', '15000-30000', 'Golang,MySQL,Redis,微服务', '无', '团队协作,沟通能力', '3年以上经验，熟悉高并发', '极高'],
        ['Java开发工程师', '负责企业级应用后端开发', '阿里巴巴', '技术', '开发', '杭州', '12000-25000', 'Java,Spring,MySQL,MyBatis', '无', '沟通能力,学习能力强', '3年以上经验', '高'],
        ['前端开发工程师', '负责Web前端开发和优化', '腾讯', '技术', '开发', '深圳', '10000-20000', 'React,Vue,TypeScript,Webpack', '无', '团队协作,创新思维', '2年以上经验', '高'],
    ]

    for row_num, data in enumerate(example_data, 2):
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.value = value

    # 设置数据行高
    for row_num in range(2, 5):
        ws.row_dimensions[row_num].height = 40

    # 添加说明工作表
    info_ws = wb.create_sheet("说明")
    info_column_widths = {'A': 15, 'B': 50}
    for col, width in info_column_widths.items():
        info_ws.column_dimensions[col].width = width

    info_headers = ['字段', '说明']
    for col_num, header in enumerate(info_headers, 1):
        cell = info_ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        cell.value = header

    info_ws.row_dimensions[1].height = 30

    info_data = [
        ['岗位名称*', '岗位名称，必填项'],
        ['描述', '岗位职责和描述'],
        ['公司', '公司名称'],
        ['行业', '如：技术、产品、运营、市场等'],
        ['类别', '如：开发、测试、设计、运维等'],
        ['地点', '工作地点，如：北京、上海、深圳等'],
        ['薪资范围', '如：10000-20000 或 面议'],
        ['技能要求', '使用逗号分隔，如：Golang,MySQL,Redis'],
        ['证书要求', '需要的证书，使用逗号分隔'],
        ['软技能', '如：团队协作、沟通能力、学习能力等'],
        ['岗位要求', '具体的岗位要求和经验要求'],
        ['成长潜力', '评估该岗位的成长空间，如：高、中、低'],
    ]

    for row_num, data in enumerate(info_data, 2):
        for col_num, value in enumerate(data, 1):
            cell = info_ws.cell(row=row_num, column=col_num)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.value = value
            info_ws.row_dimensions[row_num].height = 40

    # 保存文件
    output_path = os.path.join(os.path.dirname(__file__), '../template/job_import_template.xlsx')
    wb.save(output_path)
    print(f"✓ 岗位导入模板已生成: {output_path}")

def main():
    """主函数"""
    print("开始生成批量导入模板...")
    print()
    
    try:
        create_student_template()
        print()
        create_job_template()
        print()
        print("✓ 所有模板生成完成！")
        print()
        print("模板文件位置:")
        print("  - 学生模板: template/student_import_template.xlsx")
        print("  - 岗位模板: template/job_import_template.xlsx")
    except Exception as e:
        print(f"✗ 生成模板失败: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()